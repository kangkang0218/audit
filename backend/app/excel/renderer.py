from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.excel.contract import SHEET_HEADERS, TEMPLATE_VERSION
from app.core.privacy import mask_sensitive_text
from app.schemas.review import ReviewWorkbook

SENSITIVE_PATTERN = re.compile(r"(?<!\d)(\d{6})\d{8,11}([0-9Xx]{4})(?!\d)")
DATA_KEYS = {
    "概览": "overview",
    "基础事实提取": "facts",
    "无法提取原因": "unavailable",
    "判断建议": "findings",
    "人员社保提取": "personnel",
    "判断依据范围": "bases",
    "需补充材料": "required_materials",
}


def mask_sensitive(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    return mask_sensitive_text(SENSITIVE_PATTERN.sub(r"\1********\2", value))


def render_review_workbook(review: ReviewWorkbook, output_path: Path) -> Path:
    if review.template_version != TEMPLATE_VERSION:
        raise ValueError(f"unsupported template version: {review.template_version}")

    workbook = Workbook()
    workbook.remove(workbook.active)
    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    warning_fill = PatternFill("solid", fgColor="FCE4D6")

    payload = review.model_dump()
    for sheet_name, headers in SHEET_HEADERS.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append([sheet_name])
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        sheet.cell(1, 1).font = Font(color="FFFFFF", bold=True, size=14)
        sheet.cell(1, 1).fill = title_fill
        sheet.append(
            [
                f"模板版本：{TEMPLATE_VERSION}；默认脱敏输出；仅作风险提示与人工复核辅助。"
            ]
        )
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        sheet.append([])
        sheet.append(headers)

        for cell in sheet[4]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        rows = payload[DATA_KEYS[sheet_name]]
        for index, item in enumerate(rows, start=1):
            values = []
            for header in headers:
                if header == "序号":
                    values.append(item.get(header, index))
                else:
                    values.append(mask_sensitive(item.get(header, "")))
            sheet.append(values)
            risk_terms = ("存在疑点", "需人工复核", "当前材料不足", "证据不足", "无法")
            if item.get("是否红色标注") == "是" or any(
                term in str(item) for term in risk_terms
            ):
                for cell in sheet[sheet.max_row]:
                    cell.fill = warning_fill
                    cell.font = Font(color="9C0006")

        sheet.freeze_panes = "A5"
        sheet.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{max(4, sheet.max_row)}"
        for column_index, header in enumerate(headers, start=1):
            width = min(max(len(header) * 2 + 4, 10), 42)
            sheet.column_dimensions[get_column_letter(column_index)].width = width
        for row in sheet.iter_rows(min_row=5):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path
