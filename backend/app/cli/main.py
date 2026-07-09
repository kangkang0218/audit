from __future__ import annotations

import argparse
import hashlib
import json
import sys
from pathlib import Path
from typing import Any

import fitz
from openpyxl import load_workbook

from app.core.config import get_settings
from app.excel.contract import (
    LEGACY_HEADER_ALIASES,
    LEGACY_SHEET_ALIASES,
    SHEET_HEADERS,
    TEMPLATE_VERSION,
)
from app.excel.renderer import render_review_workbook
from app.schemas.review import ReviewWorkbook
from app.services.single_file_review import run_review
from app.llm.config_health import format_llm_config_health, get_llm_config_health
from app.llm.verification import verify_qwen, verify_qwen_vision
from app.services.page_evaluation import (
    EXTERNAL_PROCESSING_NOTICE,
    evaluate_pages,
    inspect_page_evaluation,
    parse_pages,
    parse_task_map,
)
from app.services.build_excel_with_evaluation import (
    build_excel_with_evaluation,
)
from app.services.candidate_planner import (
    inspect_candidate_plan,
    parse_material_types,
    parse_page_selection,
    plan_single_file_review,
)


def inspect_pdf(path: Path) -> dict[str, Any]:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)

    fitz.TOOLS.mupdf_display_errors(False)
    with fitz.open(path) as document:
        text_pages = 0
        low_text_pages = 0
        character_count = 0
        for page in document:
            text = page.get_text("text").strip()
            character_count += len(text)
            text_pages += bool(text)
            low_text_pages += len(text) < 20
        return {
            "source_file": path.name,
            "sha256": digest.hexdigest(),
            "file_size": path.stat().st_size,
            "page_count": len(document),
            "text_page_count": text_pages,
            "text_coverage": round(text_pages / len(document), 4) if document else 0,
            "text_character_count": character_count,
            "low_text_page_count": low_text_pages,
            "requires_ocr_candidate_pages": low_text_pages,
            "metadata_keys": sorted(key for key, value in document.metadata.items() if value),
        }


def validate_template(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    actual = {LEGACY_SHEET_ALIASES.get(name, name): name for name in workbook.sheetnames}
    missing_sheets = [name for name in SHEET_HEADERS if name not in actual]
    header_differences: dict[str, dict[str, list[str]]] = {}

    for expected_sheet, expected_headers in SHEET_HEADERS.items():
        if expected_sheet not in actual:
            continue
        sheet = workbook[actual[expected_sheet]]
        actual_headers = [
            cell.value for cell in next(sheet.iter_rows(min_row=4, max_row=4)) if cell.value
        ]
        missing_headers = [
            header
            for header in expected_headers
            if not (LEGACY_HEADER_ALIASES.get(header, {header}) & set(actual_headers))
        ]
        extra_headers = [header for header in actual_headers if header not in expected_headers]
        if missing_headers or extra_headers:
            header_differences[expected_sheet] = {
                "missing": missing_headers,
                "extra": extra_headers,
            }

    return {
        "template_version": TEMPLATE_VERSION,
        "file": path.name,
        "compatible": not missing_sheets,
        "compatibility": (
            "incompatible"
            if missing_sheets
            else ("exact" if not header_differences else "adaptable")
        ),
        "missing_sheets": missing_sheets,
        "header_differences": header_differences,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(prog="bid-review")
    subcommands = parser.add_subparsers(dest="command", required=True)

    pdf_parser = subcommands.add_parser("inspect-pdf", help="清点 PDF，不输出原始敏感文本")
    pdf_parser.add_argument("pdf", type=Path)
    pdf_parser.add_argument("--output", type=Path)

    template_parser = subcommands.add_parser("validate-template", help="校验历史 Excel 契约")
    template_parser.add_argument("xlsx", type=Path)

    render_parser = subcommands.add_parser("render-excel", help="将 review.json 渲染为固定结构 Excel")
    render_parser.add_argument("review_json", type=Path)
    render_parser.add_argument("output", type=Path)

    review_parser = subcommands.add_parser("review", help="执行单份投标 PDF 端到端审查")
    review_parser.add_argument("--input", type=Path, required=True)
    review_parser.add_argument("--output", type=Path, required=True)

    subcommands.add_parser("check-llm-config", help="仅显示脱敏的 LLM 配置健康状态")

    verify_parser = subcommands.add_parser("verify-llm", help="显式执行最小 LLM 连通性测试")
    verify_parser.add_argument("--provider", choices=["qwen"], required=True)

    vision_parser = subcommands.add_parser(
        "verify-vision",
        help="显式使用合成 PNG 执行最小视觉模型连通性测试",
    )
    vision_parser.add_argument("--provider", choices=["qwen"], required=True)

    evaluation_parser = subcommands.add_parser(
        "evaluate-pages",
        help="显式评测少量指定 PDF 实际页（会调用外部视觉模型）",
    )
    evaluation_parser.add_argument("--input", type=Path, required=True)
    evaluation_parser.add_argument("--pages", required=True)
    evaluation_parser.add_argument("--output", type=Path, required=True)
    evaluation_parser.add_argument("--provider", choices=["qwen"], required=True)
    evaluation_parser.add_argument(
        "--task-map",
        help="可选显式任务映射，如 5:bid_letter,6:bid_appendix",
    )
    evaluation_parser.add_argument(
        "--task-type",
        choices=[
            "bid_letter",
            "bid_appendix",
            "authorization",
            "guarantee",
            "personnel_table",
            "social_security",
        ],
        help="将指定页面显式作为同一种任务处理",
    )
    evaluation_parser.add_argument(
        "--retry-timeout-failures",
        action="store_true",
        help="仅允许人员表超时失败最多重试一次",
    )
    evaluation_parser.add_argument(
        "--personnel-table-strategy",
        choices=["single", "auto"],
        default="single",
        help="人员表使用单次整页或不完整时自动分块策略",
    )
    evaluation_parser.add_argument(
        "--confirm-external-processing",
        action="store_true",
        help="确认已获授权向外部模型发送指定页面",
    )

    inspect_evaluation_parser = subcommands.add_parser(
        "inspect-page-evaluation",
        help="本地检查页面评测结果，不调用模型",
    )
    inspect_evaluation_parser.add_argument("--input", type=Path, required=True)

    build_evaluation_parser = subcommands.add_parser(
        "build-excel-with-evaluation",
        help="离线导入既有页面评测结果并生成最终 Excel",
    )
    build_evaluation_parser.add_argument("--input", type=Path, required=True)
    build_evaluation_parser.add_argument(
        "--evaluation",
        type=Path,
        action="append",
        required=True,
    )
    build_evaluation_parser.add_argument("--output", type=Path, required=True)
    build_evaluation_parser.add_argument(
        "--allow-legacy-filename-only-match",
        action="store_true",
    )

    plan_parser = subcommands.add_parser(
        "plan-single-file-review",
        help="离线扫描整份 PDF 并生成有限候选页处理计划",
    )
    plan_parser.add_argument("--input", type=Path, required=True)
    plan_parser.add_argument("--output", type=Path, required=True)
    plan_parser.add_argument("--include-pages")
    plan_parser.add_argument("--exclude-pages")
    plan_parser.add_argument("--only-material-types")
    plan_parser.add_argument(
        "--max-vision-candidates",
        type=int,
        default=30,
    )

    inspect_plan_parser = subcommands.add_parser(
        "inspect-candidate-plan",
        help="本地查看既有候选页计划，不重新扫描 PDF",
    )
    inspect_plan_parser.add_argument("--input", type=Path, required=True)
    return parser


def main() -> None:
    args = build_parser().parse_args()
    try:
        if args.command == "inspect-pdf":
            result = inspect_pdf(args.pdf)
            output = json.dumps(result, ensure_ascii=False, indent=2)
            if args.output:
                args.output.write_text(output + "\n", encoding="utf-8")
            else:
                print(output)
        elif args.command == "validate-template":
            result = validate_template(args.xlsx)
            print(json.dumps(result, ensure_ascii=False, indent=2))
            if not result["compatible"]:
                sys.exit(2)
        elif args.command == "render-excel":
            payload = json.loads(args.review_json.read_text(encoding="utf-8"))
            review = ReviewWorkbook.model_validate(payload)
            render_review_workbook(review, args.output)
            print(args.output)
        elif args.command == "review":
            paths = run_review(args.input, args.output, get_settings())
            print(json.dumps({key: str(path) for key, path in paths.items()}, ensure_ascii=False, indent=2))
        elif args.command == "check-llm-config":
            health = get_llm_config_health(get_settings())
            print(format_llm_config_health(health))
        elif args.command == "verify-llm":
            result = verify_qwen(get_settings())
            print(f"Provider reachable: {str(result.reachable).lower()}")
            print(f"JSON schema validated: {str(result.schema_validated).lower()}")
            if result.reachable:
                print(f"Model name: {result.model_name}")
            else:
                if result.http_status is not None:
                    print(f"HTTP status: {result.http_status}")
                print(f"Error type: {result.error_type}")
                print(f"Error summary: {result.error_summary}")
                sys.exit(2)
        elif args.command == "verify-vision":
            result = verify_qwen_vision(get_settings())
            print(f"Provider reachable: {str(result.reachable).lower()}")
            print(f"Vision request accepted: {str(result.request_accepted).lower()}")
            print(f"JSON schema validated: {str(result.schema_validated).lower()}")
            if result.reachable:
                print(f"Model name: {result.model_name}")
                print(f"Elapsed seconds: {result.elapsed_seconds}")
            else:
                if result.http_status is not None:
                    print(f"HTTP status: {result.http_status}")
                print(f"Error type: {result.error_type}")
                print(f"Error summary: {result.error_summary}")
                sys.exit(2)
        elif args.command == "evaluate-pages":
            print(EXTERNAL_PROCESSING_NOTICE)
            if not args.confirm_external_processing:
                print(
                    "\n未执行模型调用。确认已获授权后，请显式运行：\n"
                    f"{sys.executable} -m app.cli.main evaluate-pages "
                    f'--input "{args.input}" --pages "{args.pages}" '
                    f'--output "{args.output}" --provider {args.provider} '
                    "--confirm-external-processing",
                    file=sys.stderr,
                )
                sys.exit(2)
            summary = evaluate_pages(
                args.input,
                parse_pages(args.pages),
                args.output,
                get_settings(),
                provider_name=args.provider,
                task_map=parse_task_map(args.task_map),
                task_type=args.task_type,
                retry_timeout_failures=args.retry_timeout_failures,
                personnel_table_strategy=args.personnel_table_strategy,
            )
            print(summary.model_dump_json(indent=2))
        elif args.command == "inspect-page-evaluation":
            report = inspect_page_evaluation(args.input)
            print(json.dumps(report, ensure_ascii=False, indent=2))
        elif args.command == "build-excel-with-evaluation":
            result = build_excel_with_evaluation(
                args.input,
                args.evaluation,
                args.output,
                allow_legacy_filename_only_match=(
                    args.allow_legacy_filename_only_match
                ),
            )
            print(
                json.dumps(
                    {
                        "output_files": {
                            key: str(path)
                            for key, path in result["paths"].items()
                        },
                        "imported_results": len(
                            result["imported_results"]
                        ),
                        "conflict_count": result["import_audit"][
                            "conflict_count"
                        ],
                        "legacy_match_methods": result["import_audit"][
                            "legacy_match_methods"
                        ],
                        "external_model_calls": 0,
                        "ocr_calls": 0,
                        "http_calls": 0,
                    },
                    ensure_ascii=False,
                    indent=2,
                )
            )
        elif args.command == "plan-single-file-review":
            result = plan_single_file_review(
                args.input,
                args.output,
                include_pages=parse_page_selection(args.include_pages),
                exclude_pages=parse_page_selection(args.exclude_pages),
                only_material_types=parse_material_types(
                    args.only_material_types
                ),
                max_vision_candidates=args.max_vision_candidates,
            )
            print(
                json.dumps(
                    {
                        "output_files": {
                            key: str(path)
                            for key, path in result["paths"].items()
                        },
                        "processing_summary": result[
                            "processing_summary"
                        ],
                        "candidate_ranges": len(result["ranges"]),
                        "external_model_calls": 0,
                        "ocr_calls": 0,
                        "http_calls": 0,
                    },
                    ensure_ascii=False,
                    indent=2,
                )
            )
        elif args.command == "inspect-candidate-plan":
            result = inspect_candidate_plan(args.input)
            print(json.dumps(result, ensure_ascii=False, indent=2))
    except (OSError, ValueError, json.JSONDecodeError) as exc:
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
