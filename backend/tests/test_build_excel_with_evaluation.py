from __future__ import annotations

import json
from pathlib import Path

import httpx
from openpyxl import load_workbook

from app.excel.contract import SHEET_HEADERS
from app.services.build_excel_with_evaluation import (
    build_excel_with_evaluation,
)
from tests.evaluation_test_support import make_evaluation, make_pdf


def test_offline_build_generates_real_seven_sheet_excel(
    tmp_path: Path,
    monkeypatch,
) -> None:
    pdf = tmp_path / "sample.pdf"
    make_pdf(
        pdf,
        [
            "（一）投标函 项目名称 测试项目 投标人：测试有限公司",
        ],
    )
    evaluation = tmp_path / "evaluation"
    make_evaluation(evaluation, pdf)
    output = tmp_path / "output"

    def forbidden(*args, **kwargs):
        raise AssertionError("离线流程不得初始化 Provider、HTTP 或 OCR")

    monkeypatch.setattr(httpx, "post", forbidden)
    monkeypatch.setattr(httpx, "request", forbidden)
    monkeypatch.setattr(
        "app.llm.qwen_provider.QwenVisionProvider.__init__",
        forbidden,
    )
    monkeypatch.setattr(
        "app.llm.qwen_provider.QwenTextProvider.__init__",
        forbidden,
    )
    monkeypatch.setattr(
        "app.parsers.ocr_parser.OptionalPaddleOCR.__init__",
        forbidden,
    )

    result = build_excel_with_evaluation(
        pdf,
        [evaluation],
        output,
    )

    required = {
        "pages.json",
        "toc.json",
        "facts.json",
        "personnel.json",
        "findings.json",
        "review.json",
        "evaluation_import_audit.json",
        "field_provenance.json",
        "audit_log.json",
    }
    assert required <= {path.name for path in output.iterdir()}
    workbook = load_workbook(result["paths"]["excel"])
    assert workbook.sheetnames == list(SHEET_HEADERS)
    assert all(sheet.max_row > 4 for sheet in workbook.worksheets)
    facts = json.loads((output / "facts.json").read_text())
    project = next(item for item in facts if item["field_name"] == "项目名称")
    assert project["merge_status"] == "qwen_only"
    audit = json.loads((output / "audit_log.json").read_text())
    assert audit[-1]["details"] == {
        "external_model_calls": 0,
        "ocr_calls": 0,
        "http_calls": 0,
    }
