import json
import re

from openpyxl import load_workbook

from app.excel.contract import SHEET_HEADERS
from app.validators.single_file import SINGLE_FILE_BOUNDARIES


def test_real_pdf_end_to_end_outputs(end_to_end_output) -> None:
    required = {
        "pages.json",
        "toc.json",
        "facts.json",
        "personnel.json",
        "personnel_extraction_audit.json",
        "rejected_personnel_candidates.json",
        "findings.json",
        "review.json",
        "audit_log.json",
        "llm_calls.jsonl",
    }
    assert required <= {path.name for path in end_to_end_output.iterdir()}

    facts = {
        item["field_name"]: item
        for item in json.loads((end_to_end_output / "facts.json").read_text())
    }
    people = json.loads((end_to_end_output / "personnel.json").read_text())
    findings = json.loads((end_to_end_output / "findings.json").read_text())
    assert facts["投标人名称"]["value"] == "中和刚大工程顾问有限公司"
    assert facts["项目负责人"]["value"] == "张军校"
    assert facts["基本费率折扣系数"]["value"] == "0.53"
    assert facts["审减费率折扣系数"]["value"] == "0.53"
    assert facts["保证金声明金额"]["value"] == "10000 元"
    assert {"张军校", "杨敬宇", "杜娟", "盖泰嘉", "李杨"} <= {
        item["name"] for item in people
    }
    assert set(SINGLE_FILE_BOUNDARIES) <= {item["result"] for item in findings}

    excel = next(end_to_end_output.glob("*.xlsx"))
    assert load_workbook(excel, read_only=True).sheetnames == list(SHEET_HEADERS)


def test_outputs_do_not_contain_full_sensitive_numbers(end_to_end_output) -> None:
    sensitive = re.compile(
        r"(?<!\d)(?:\d{17}[0-9Xx]|1[3-9]\d{9}|\d{16,24})(?!\d)"
    )
    for path in end_to_end_output.glob("*.json*"):
        assert sensitive.search(path.read_text(encoding="utf-8")) is None, path.name
