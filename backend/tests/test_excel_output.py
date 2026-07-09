from openpyxl import load_workbook

from app.excel.contract import SHEET_HEADERS


def test_real_excel_has_complete_structure_and_risk_style(end_to_end_output) -> None:
    excel = next(end_to_end_output.glob("*.xlsx"))
    workbook = load_workbook(excel)

    assert workbook.sheetnames == list(SHEET_HEADERS)
    for sheet in workbook.worksheets:
        assert sheet.max_row > 4
        assert sheet.freeze_panes == "A5"
        assert sheet.auto_filter.ref
    finding_sheet = workbook["判断建议"]
    assert any(
        cell.fill.fgColor.rgb in {"00FCE4D6", "FCE4D6"}
        for row in finding_sheet.iter_rows(min_row=5)
        for cell in row
    )

