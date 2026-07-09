import asyncio
from io import BytesIO

import fitz
import httpx
from openpyxl import load_workbook

from app.main import app


def _pdf_bytes(page_count: int = 2) -> bytes:
    document = fitz.open()
    for index in range(page_count):
        page = document.new_page()
        page.insert_text((72, 72), f"Bid review test page {index + 1}")
    content = document.tobytes()
    document.close()
    return content


async def _request(
    method: str,
    path: str,
    **kwargs,
) -> httpx.Response:
    transport = httpx.ASGITransport(app=app)
    async with httpx.AsyncClient(
        transport=transport,
        base_url="http://test",
        timeout=30,
    ) as client:
        return await client.request(method, path, **kwargs)


def test_frontend_is_served_by_current_backend() -> None:
    response = asyncio.run(_request("GET", "/"))

    assert response.status_code == 200
    assert "投标文件三项审查" in response.text
    assert "20260706-1" in response.text


def test_pdf_info_uses_current_api_prefix() -> None:
    response = asyncio.run(
        _request(
            "POST",
            "/api/v1/pdf-info",
            files={"pdf": ("sample.pdf", _pdf_bytes(2), "application/pdf")},
        )
    )

    assert response.status_code == 200
    assert response.json() == {"pageCount": 2}


def test_bid_review_returns_real_seven_sheet_workbook() -> None:
    response = asyncio.run(
        _request(
            "POST",
            "/api/v1/bid-review",
            data={"mode": "bid_review_three_items"},
            files={"pdf": ("sample.pdf", _pdf_bytes(1), "application/pdf")},
        )
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["x-review-mode"] == "local-only"
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    assert len(workbook.sheetnames) == 7


def test_upload_rejects_non_pdf_content() -> None:
    response = asyncio.run(
        _request(
            "POST",
            "/api/v1/pdf-info",
            files={"pdf": ("fake.pdf", b"not-a-pdf", "application/pdf")},
        )
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "文件内容不是有效的 PDF。"
