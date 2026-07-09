from pathlib import Path

from openpyxl import load_workbook

from app.excel.contract import SHEET_HEADERS
from app.excel.renderer import mask_sensitive, render_review_workbook
from app.schemas.review import ReviewWorkbook


def test_mask_sensitive_number() -> None:
    assert mask_sensitive("身份证 410422199706260048") == "身份证 410422********0048"


def test_render_fixed_workbook(tmp_path: Path) -> None:
    review = ReviewWorkbook(
        source_file="sample.pdf",
        overview=[{"项目": "源文件", "内容": "sample.pdf"}],
        personnel=[
            {
                "姓名": "张某",
                "备注": "身份证 410422199706260048，需人工复核",
            }
        ],
    )
    output = render_review_workbook(review, tmp_path / "result.xlsx")
    workbook = load_workbook(output, read_only=True)

    assert workbook.sheetnames == list(SHEET_HEADERS)
    assert workbook["人员社保提取"]["O5"].value == "身份证 410422********0048，需人工复核"

